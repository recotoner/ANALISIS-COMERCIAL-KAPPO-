from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SHEET_NAMES = [
    "Resumen",
    "Guia comercial",
    "Revision prioritaria",
    "Seguimiento sugerido",
    "Oportunidades",
    "Inactivos relevantes",
    "Inteligencia tiempo comercial",
    "Ranking clientes",
]

AGENDA_COLUMNS = {
    "cliente": "Cliente",
    "vendedor_responsable_sugerido": "Vendedor responsable sugerido",
    "venta_total": "Venta total",
    "participacion": "Participacion %",
    "prioridad_gestion": "Prioridad sugerida",
    "tipo_sugerencia": "Tipo de sugerencia",
    "alerta_comercial": "Alerta comercial",
    "categoria_recompra": "Categoria de recompra",
    "lectura_comercial_patron": "Lectura comercial del patron",
    "sugerencia_por_tiempo": "Sugerencia por tiempo",
    "ultima_compra": "Ultima compra",
    "ultimo_mes_con_compra": "Ultimo mes con compra",
    "dias_desde_ultima_compra": "Dias desde ultima compra",
    "proxima_compra_esperada": "Proxima compra esperada",
    "dias_atraso": "Dias de atraso",
    "intervalo_mensual_mediano_dias": "Ciclo tipico entre meses con compra",
    "confianza_recompra": "Confianza",
    "motivo_comercial": "Motivo comercial",
    "validacion_requerida_crm": "Validacion requerida en CRM",
    "venta_ultimos_3_meses": "Venta ultimos 3 meses",
    "venta_3_meses_anteriores": "Venta 3 meses anteriores",
    "variacion_absoluta": "Variacion absoluta",
    "variacion_porcentual": "Variacion %",
    "ultimo_mes_con_compra_alertas": "Ultimo mes con compra alertas",
    "tendencia": "Tendencia",
    "recomendacion_sugerida": "Recomendacion sugerida",
}

TIME_INTELLIGENCE_COLUMNS = {
    "cliente": "Cliente",
    "vendedor_responsable_sugerido": "Vendedor responsable sugerido",
    "ultima_compra": "Ultima compra",
    "ultimo_mes_con_compra": "Ultimo mes con compra",
    "dias_desde_ultima_compra": "Dias desde ultima compra",
    "categoria_recompra": "Categoria de recompra",
    "lectura_comercial_patron": "Lectura comercial del patron",
    "sugerencia_por_tiempo": "Sugerencia por tiempo",
    "proxima_compra_esperada": "Proxima compra esperada",
    "dias_atraso": "Dias de atraso",
    "intervalo_mensual_mediano_dias": "Ciclo tipico entre meses con compra",
    "confianza_recompra": "Confianza",
    "venta_total": "Venta total",
    "participacion": "Participacion %",
}

RANKING_COLUMNS = {
    "ranking": "Ranking",
    "cliente": "Cliente",
    "venta_total": "Venta total",
    "participacion": "Participacion %",
    "participacion_acumulada": "Participacion acumulada %",
    "clasificacion_cliente": "Clasificacion cliente",
}

MONEY_COLUMNS = {
    "Venta total",
    "Venta ultimos 3 meses",
    "Venta 3 meses anteriores",
    "Variacion absoluta",
}
PERCENTAGE_COLUMNS = {
    "Participacion %",
    "Participacion acumulada %",
    "Variacion %",
}
DATE_COLUMNS = {"Ultima compra", "Proxima compra esperada"}
METHODOLOGY_NOTE = (
    "La recompra esperada se calcula considerando meses calendario con compra, "
    "no documentos individuales dentro del mismo mes. Es una referencia de "
    "gestion comercial y no una prediccion exacta. Cuando existe historial "
    "insuficiente, compras concentradas o patron irregular, el sistema lo "
    "indica para evitar falsas alertas."
)


def build_commercial_management_excel(
    summary: pd.DataFrame,
    agenda: pd.DataFrame,
    client_alerts: pd.DataFrame,
    ranking: pd.DataFrame,
    commercial_time_clients: pd.DataFrame | None = None,
) -> bytes:
    agenda = agenda.copy()
    client_alerts = client_alerts.copy()
    commercial_time_clients = (
        pd.DataFrame()
        if commercial_time_clients is None
        else commercial_time_clients.copy()
    )

    review_priority = _filter_equals(agenda, "prioridad_gestion", "Alta")
    suggested_follow_up = _filter_equals(agenda, "prioridad_gestion", "Media")
    if "alerta_comercial" in suggested_follow_up.columns:
        suggested_follow_up = suggested_follow_up[
            suggested_follow_up["alerta_comercial"] != "Oportunidad"
        ]
    opportunities = _filter_equals(agenda, "alerta_comercial", "Oportunidad")
    relevant_inactive = _filter_equals(agenda, "tendencia", "Inactivo")
    if relevant_inactive.empty:
        relevant_inactive = _filter_equals(client_alerts, "tendencia", "Inactivo")
    if "venta_total" in relevant_inactive.columns:
        relevant_inactive = relevant_inactive.sort_values(
            "venta_total", ascending=False, kind="stable"
        )

    sheets = {
        "Resumen": _with_methodology_note(summary),
        "Guia comercial": _prepare_table(agenda, AGENDA_COLUMNS),
        "Revision prioritaria": _prepare_table(review_priority, AGENDA_COLUMNS),
        "Seguimiento sugerido": _prepare_table(suggested_follow_up, AGENDA_COLUMNS),
        "Oportunidades": _prepare_table(opportunities, AGENDA_COLUMNS),
        "Inactivos relevantes": _prepare_table(relevant_inactive, AGENDA_COLUMNS),
        "Inteligencia tiempo comercial": _prepare_table(
            commercial_time_clients,
            TIME_INTELLIGENCE_COLUMNS,
        ),
        "Ranking clientes": _prepare_table(ranking, RANKING_COLUMNS),
    }

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name in SHEET_NAMES:
            sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
            _style_worksheet(writer.book[sheet_name])

    return output.getvalue()


def _prepare_table(data: pd.DataFrame, column_map: dict[str, str]) -> pd.DataFrame:
    available_columns = [column for column in column_map if column in data.columns]
    table = data.loc[:, available_columns].rename(columns=column_map).copy()
    return _format_missing_commercial_time_values(table)


def _format_missing_commercial_time_values(table: pd.DataFrame) -> pd.DataFrame:
    if "Proxima compra esperada" in table.columns:
        table["Proxima compra esperada"] = table["Proxima compra esperada"].where(
            table["Proxima compra esperada"].notna(),
            "No determinada",
        )
    if "Ciclo tipico entre meses con compra" in table.columns:
        table["Ciclo tipico entre meses con compra"] = table[
            "Ciclo tipico entre meses con compra"
        ].where(
            table["Ciclo tipico entre meses con compra"].notna(),
            "N/D",
        )
    return table


def _with_methodology_note(summary: pd.DataFrame) -> pd.DataFrame:
    summary = summary.copy()
    if {"Indicador", "Valor"}.issubset(summary.columns):
        return pd.concat(
            [
                summary,
                pd.DataFrame(
                    [
                        {
                            "Indicador": "Nota metodologica recompra",
                            "Valor": METHODOLOGY_NOTE,
                        }
                    ]
                ),
            ],
            ignore_index=True,
        )
    summary["Nota metodologica recompra"] = METHODOLOGY_NOTE
    return summary


def _filter_equals(data: pd.DataFrame, column: str, value: str) -> pd.DataFrame:
    if column not in data.columns:
        return data.iloc[0:0].copy()
    return data[data[column] == value].copy()


def _style_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="082D5F")
    header_font = Font(color="FFFFFF", bold=True)

    worksheet.freeze_panes = "A2"
    if worksheet.max_column and worksheet.max_row:
        worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 24

    for column_index, cells in enumerate(worksheet.iter_cols(), start=1):
        header = cells[0].value
        max_length = max(
            (len(str(cell.value)) for cell in cells if cell.value is not None),
            default=0,
        )
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 12),
            55,
        )
        if header in MONEY_COLUMNS:
            for cell in cells[1:]:
                cell.number_format = '$#,##0;[Red]-$#,##0'
        elif header in PERCENTAGE_COLUMNS:
            for cell in cells[1:]:
                cell.number_format = "0.0%"
        elif header in DATE_COLUMNS:
            for cell in cells[1:]:
                if not isinstance(cell.value, str):
                    cell.number_format = "DD/MM/YYYY"
