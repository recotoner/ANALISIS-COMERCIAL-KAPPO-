import unittest
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from src.commercial_excel import SHEET_NAMES, build_commercial_management_excel


class CommercialExcelTests(unittest.TestCase):
    def test_builds_expected_sheets_and_segments(self):
        summary = pd.DataFrame([{"Indicador": "Venta total", "Valor": "$1.000"}])
        agenda = pd.DataFrame(
            [
                {
                    "cliente": "Cliente critico",
                    "vendedor_responsable_sugerido": "Ana",
                    "categoria_recompra": "Recompra esperada atrasada",
                    "lectura_comercial_patron": "Ciclo consistente",
                    "sugerencia_por_tiempo": "Revisar recuperacion comercial.",
                    "proxima_compra_esperada": pd.Timestamp("2026-07-01"),
                    "dias_atraso": 0,
                    "intervalo_mensual_mediano_dias": 31,
                    "confianza_recompra": "Alta",
                    "venta_total": 1000,
                    "participacion": 0.50,
                    "tendencia": "En caida",
                    "alerta_comercial": "Alerta critica",
                    "prioridad_gestion": "Alta",
                    "tipo_sugerencia": "Revisar cliente critico",
                    "validacion_requerida_crm": "Revisar CRM",
                    "motivo_comercial": "Caida reciente",
                },
                {
                    "cliente": "Cliente oportunidad",
                    "vendedor_responsable_sugerido": "Luis",
                    "venta_total": 500,
                    "participacion": 0.25,
                    "tendencia": "En crecimiento",
                    "alerta_comercial": "Oportunidad",
                    "prioridad_gestion": "Media",
                    "tipo_sugerencia": "Evaluar desarrollo",
                    "validacion_requerida_crm": "Revisar CRM",
                    "motivo_comercial": "Crecimiento reciente",
                },
            ]
        )
        alerts = pd.DataFrame(
            [
                {
                    "cliente": "Cliente inactivo",
                    "venta_total": 800,
                    "participacion": 0.10,
                    "tendencia": "Inactivo",
                    "alerta_comercial": "Alerta relevante",
                    "recomendacion_sugerida": "Contactar",
                }
            ]
        )
        ranking = pd.DataFrame(
            [
                {
                    "ranking": 1,
                    "cliente": "Cliente critico",
                    "venta_total": 1000,
                    "participacion": 0.50,
                    "participacion_acumulada": 0.50,
                    "clasificacion_cliente": "Cliente estrategico",
                }
            ]
        )
        commercial_time = pd.DataFrame(
            [
                {
                    "cliente": "Cliente critico",
                    "vendedor_responsable_sugerido": "Ana",
                    "ultima_compra": pd.Timestamp("2026-06-01"),
                    "ultimo_mes_con_compra": "06/2026",
                    "dias_desde_ultima_compra": 30,
                    "categoria_recompra": "Recompra esperada atrasada",
                    "lectura_comercial_patron": "Ciclo consistente",
                    "sugerencia_por_tiempo": "Revisar recuperacion comercial.",
                    "proxima_compra_esperada": pd.Timestamp("2026-07-01"),
                    "dias_atraso": 0,
                    "intervalo_mensual_mediano_dias": 31,
                    "confianza_recompra": "Alta",
                    "venta_total": 1000,
                    "participacion": 0.50,
                }
            ]
        )

        content = build_commercial_management_excel(
            summary,
            agenda,
            alerts,
            ranking,
            commercial_time,
        )
        workbook = load_workbook(BytesIO(content), data_only=True)
        guide_sheet = SHEET_NAMES[1]
        review_sheet = SHEET_NAMES[2]
        opportunities_sheet = SHEET_NAMES[4]
        inactive_sheet = SHEET_NAMES[5]
        time_sheet = SHEET_NAMES[6]

        self.assertEqual(workbook.sheetnames, SHEET_NAMES)
        self.assertEqual(workbook[review_sheet].max_row, 2)
        self.assertEqual(workbook[opportunities_sheet].max_row, 2)
        self.assertEqual(workbook[guide_sheet]["C1"].value, "Venta total")
        self.assertEqual(workbook[guide_sheet]["H1"].value, "Categoria de recompra")
        self.assertEqual(
            workbook[guide_sheet]["H2"].value,
            "Recompra esperada atrasada",
        )
        self.assertEqual(workbook[guide_sheet]["I2"].value, "Ciclo consistente")
        self.assertEqual(
            workbook[guide_sheet]["J2"].value,
            "Revisar recuperacion comercial.",
        )
        self.assertEqual(
            workbook[guide_sheet]["K2"].value,
            pd.Timestamp("2026-07-01").to_pydatetime(),
        )
        self.assertEqual(workbook[guide_sheet]["L2"].value, 0)
        self.assertEqual(workbook[guide_sheet]["M2"].value, 31)
        self.assertGreaterEqual(workbook[guide_sheet]["M2"].value, 28)
        self.assertEqual(workbook[guide_sheet]["N2"].value, "Alta")
        self.assertEqual(workbook[time_sheet]["F1"].value, "Categoria de recompra")
        self.assertEqual(
            workbook[time_sheet]["F2"].value,
            "Recompra esperada atrasada",
        )
        self.assertEqual(workbook[time_sheet]["K2"].value, 31)
        self.assertEqual(
            workbook[inactive_sheet]["A2"].value,
            "Cliente inactivo",
        )
        self.assertEqual(workbook["Ranking clientes"]["C2"].value, 1000)
        self.assertEqual(workbook["Ranking clientes"]["D2"].number_format, "0.0%")


if __name__ == "__main__":
    unittest.main()
